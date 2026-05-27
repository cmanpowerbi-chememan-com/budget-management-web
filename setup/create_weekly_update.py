"""
Weekly update: merge local task ROWS with SharePoint version, then upload.
Flow:
  1. Download current SharePoint file
  2. Read existing rows (incl. colleague tasks)
  3. Upsert local ROWS (by Action name) — preserve unmatched colleague rows
  4. Build styled Excel from merged rows
  5. Save locally + upload to SharePoint

Usage: python setup/create_weekly_update.py
"""
import os
import io
import msal
import requests
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ── constants ────────────────────────────────────────────────────────────────
FOLDER = Path(__file__).parent.parent / "weekly_update"
FILENAME = "budget management web.xlsx"
SHEET_NAME = "action"

SHAREPOINT_SITE = "chememan.sharepoint.com"
SHAREPOINT_TEAM = "CMANDigitalTechnology"
SHAREPOINT_FOLDER = "General/05 Data Analytics/03 Project/6.Budgeting and Management"

TENANT_ID = os.getenv("ENTRA_TENANT_ID")
CLIENT_ID = os.getenv("ENTRA_CLIENT_ID")
CLIENT_SECRET = os.getenv("ENTRA_CLIENT_SECRET")

# ── color palette ────────────────────────────────────────────────────────────
C = {
    "header_bg":  "1F1F1F",
    "header_txt": "FFD700",
    "proj_bg":    "4472C4",
    "pipe_bg":    "2E75B6",
    "dev_bg":     "375623",
    "action_bg":  "1F3864",
    "done_bg":    "70AD47",
    "prog_bg":    "F4B183",
    "white":      "FFFFFF",
    "black":      "000000",
    "row_alt":    "EBF3FB",
    "row_norm":   "FFFFFF",
}

TYPE_COLORS = {
    "Project Settings": C["proj_bg"],
    "Data Pipeline":    C["pipe_bg"],
    "Development":      C["dev_bg"],
    "Project Action":   C["action_bg"],
}

# ── my current task rows (upserted into SharePoint version) ──────────────────
# Tuple: (Action type, Action, Remark, PIC, Cut-off, Status)
ROWS = [
    ("Project Settings", "Requirements & Architecture",
     "Tech stack finalized: React + FastAPI + Azure SQL + Fabric Lakehouse",
     "Jakkarit", "17-Apr-26", "Done"),
    ("Project Settings", "Azure SQL DB Setup",
     "DB server created, firewall configured, 5 tables created (budget_submissions, approval_status, etc.)",
     "Jakkarit", "24-Apr-26", "Done"),
    ("Project Settings", "Azure Container Registry",
     "cmanbudgetacr.azurecr.io — Basic tier, Southeast Asia",
     "Jakkarit", "24-Apr-26", "Done"),
    ("Project Settings", "Azure Container Apps",
     "cman-budget-mngt-web created with Log Analytics workspace",
     "Jakkarit", "24-Apr-26", "Done"),
    ("Project Settings", "Fabric Workspace + Lakehouse + Notebooks",
     "Workspace: budget_management_web — nb_upload_actuals, nb_send_email created",
     "Jakkarit", "24-Apr-26", "Done"),

    ("Data Pipeline", "mas_employee_data Sync Script",
     "C-POP HR API → Azure SQL, incremental sync (setup/sync_employees.py), 621 Active employees",
     "Jakkarit", "5-May-26", "Done"),
    ("Data Pipeline", "GitHub Actions Daily Sync",
     "06:00 Bangkok daily (.github/workflows/sync_employees.yml), first run verified 2026-05-05",
     "Jakkarit", "5-May-26", "Done"),
    ("Data Pipeline", "dim_cost_center Analysis",
     "Source: docs/Cost center (Update 18 Mar 26) 1.xlsx — 210 rows, orgcode join strategy identified",
     "Jakkarit", "6-May-26", "Done"),
    ("Data Pipeline", "Verify orgcode → Cost Center Link",
     "Confirmed: orgcode (C-POP format e.g. 1165403) != SAP CC format (e.g. 10IT011100) — bridge table dim_orgcode_costcenter_map required, data from Chatdanai API",
     "Jakkarit", "15-May-26", "Done"),
    ("Data Pipeline", "sap_m_cost_center Table",
     "Cost center reference table already exists in system as sap_m_cost_center — no need to create dim_cost_center",
     "Jakkarit", "6-May-26", "Done"),

    ("Development", "Weekly Update Skill",
     "Created skill/skill_weekly_update.md + upgraded create_weekly_update.py to merge SP tasks before upload",
     "Jakkarit", "6-May-26", "Done"),
    ("Development", "Project Folder Restructure",
     "Master table editors use SWA + Functions + Vanilla JS. Main budget app (submit/approve/dashboard) will use FastAPI + React — more complex state and workflow",
     "Jakkarit", "TBC", "In Progress"),
    ("Development", "FastAPI Backend",
     "For main budget app: /api/auth, /api/budget, /api/approval, /api/dashboard, /api/admin",
     "Jakkarit", "TBC", "In Progress"),
    ("Development", "React + Vite Frontend Scaffold",
     "For main budget app: Login, Submit Budget (Template 1.1), Approve (VP/Staff/Manager), Dashboard, Admin",
     "Jakkarit", "TBC", "In Progress"),

    ("Data Pipeline", "SAP Gap Columns — Ticket to SAP Team",
     "4 cols added to sap_t_gl_trans: Entry Date, Order, Object Class, Order Short Text. Ratima verified data in SAP_T_GL_TRANS_1000_RATIMA_TEST1 ✅",
     "Ratima", "26-May-26", "Done"),
    ("Data Pipeline", "SAP_M_INTERNAL_ORDER",
     "Cancelled — user not use",
     "Ratima", "13-May-26", "Done"),
    ("Data Pipeline", "SAP Job Run Setup",
     "If SAP_M_INTERNAL_ORDER + SAP_T_GL_TRANS_1000_RATIMA_TEST1 test pass, data OK → Jakkarit แจ้งกลับ SAP Team เพื่อตั้ง Scheduled Job Run ใน SAP Production",
     "Jakkarit", "13-May-26", "In Progress"),

    ("Project Action", "1st Recap Project with Laddawan",
     "นัด recap ความคืบหน้าโปรเจกต์ครั้งแรก Thu 14/05/2026 8:30-9:00 AM — ยิง MS Teams invite แล้ว",
     "Jakkarit", "14-May-26", "Done"),

    ("Project Action", "Meeting with User — Template Filling Data",
     "Discuss how users fill budget template data, confirm workflow and UX requirements",
     "Jakkarit", "13-May-26", "Done"),

    ("Project Action", "Meeting with User — Special Template Phase 2",
     "Discuss special template requirements (phase 2) — confirm workflow and UX for special cases",
     "Jakkarit", "TBC", "In Progress"),

    ("Data Pipeline", "Cost Center -> Line Officer Link (Chatdanai API)",
     "orgcode → cost_center mapping received. Source: docs/09orgcode & costcenter.xlsx (729 rows, 183 orgcodes, 205 CCs) — many-to-many confirmed",
     "Chatdanai", "26-May-26", "Done"),

    ("Data Pipeline", "OneLake Connection",
     "cman-fabric-write added to workspace, setup/connect_onelake.py created, connection verified",
     "Jakkarit", "07-May-26", "Done"),

    ("Data Pipeline", "On-premises Gateway",
     "cman-dw-dev-conn-gateway created on E-SMARTISO server, Fabric Data Pipeline: D:/SAPDW/PRD -> Files/00landing configured",
     "Jakkarit", "07-May-26", "Done"),

    ("Data Pipeline", "00landing Files",
     "3 SAP test files landed: SAP_T_GL_TRANS_1000_RATIMA_TEST1 (233,372 rows/116MB), SAP_M_COST_CENTER (341), SAP_M_INTERNAL_ORDER (2,543)",
     "Jakkarit", "07-May-26", "Done"),

    ("Data Pipeline", "nb_landing_to_bronze",
     "bronze_ACDOCA (233,372), bronze_sap_m_cost_center (341), bronze_sap_m_internal_order (2,543) loaded as Delta tables",
     "Jakkarit", "07-May-26", "Done"),

    ("Data Pipeline", "nb_bronze_to_silver",
     "silver_sap_gl_trans (88 cols, sign-flip on 6 amount/FX cols), silver_cost_center_master, silver_internal_order_master",
     "Jakkarit", "07-May-26", "Done"),

    ("Data Pipeline", "entry_date null issue",
     "SAP date format = yyyyMMdd (no dash) — fixed sc() helper to use to_date(c, 'yyyyMMdd'), covers all date cols automatically, entry_date verified not null",
     "Jakkarit", "07-May-26", "Done"),

    ("Data Pipeline", "Bronze->Silver Data Consistency Check",
     "Verified: sap_m_gl_account row count match, non-null counts match, distinct counts match, date cols parse correctly ✅",
     "Jakkarit", "26-May-26", "Done"),

    ("Data Pipeline", "Cost Type Master — cost_type & cost_sub_type",
     "Added cost_type + cost_sub_type cols to docs/cost type master.xlsx: GL5→Indirect, GL6+10SC012000→Logistic/Admin, GL6+10CM*/10SC*(≠10SC012000)→SGA/Selling, GL6 others→SGA/Admin",
     "Jakkarit", "13-May-26", "Done"),

    ("Development", "Budget Form UX Design Analysis",
     "CC hierarchy verified (210 CC→109 dept→38 div). 1 user=1 dept fills all CCs. GL dropdown by actuals history per CC (not 52/62 prefix rule). User does 3 things: select CC, select GL, enter Jan-Dec.",
     "Jakkarit", "TBC", "In Progress"),

    ("Data Pipeline", "dim_costcenter_division_map",
     "Create table in Azure SQL, load 210 rows from 02cost center & department (master).xlsx — cost_center→division+department mapping",
     "Jakkarit", "TBC", "In Progress"),

    ("Data Pipeline", "dim_orgcode_costcenter_map",
     "Create table in Azure SQL — orgcode→cost_center bridge. Source: file from DW project (same source Chatdanai's team uses), not API. Need file path from DW project.",
     "Jakkarit", "TBC", "In Progress"),

    ("Data Pipeline", "dim_gl_master",
     "Create table in Azure SQL, load 137 rows from 04gl code & gl group & gl thai name (master).xlsx — gl_account→group_exp+thai_name",
     "Jakkarit", "TBC", "In Progress"),

    ("Development", "Landing Page Design",
     "Design step: landing page based on docs/index.html (Chememan branding, dark/light theme, glassmorphism) — not yet built",
     "Jakkarit", "TBC", "In Progress"),

    ("Development", "Module 1 — Import/Export Approved Budget",
     "Design step: budget dept export approved_budget_{yyyy}.csv (15 cols), import with validation, overwrite-by-year, near-realtime table — not yet built",
     "Jakkarit", "TBC", "In Progress"),

    ("Development", "0003-gl-group Master Table Editor",
     "GL Group CRUD on Azure SWA — GL codes from Fabric Lakehouse gold_sap_m_gl_account_group_name (430 SAP codes), sort/filter/export CSV, event delegation CSP compliance, thread-local DB connections. Deployed + tested ✅",
     "Jakkarit", "26-May-26", "Done"),

    ("Data Pipeline", "silver_sap_m_gl_account (GL Account Master)",
     "bronze_sap_m_gl_account → silver via sc() helper, ERDAT cast with to_date(c,'yyyyMMdd'), consistency check done",
     "Jakkarit", "26-May-26", "Done"),

    ("Data Pipeline", "gold_sap_m_gl_account_group_name",
     "gold_sap_m_gl_account (filter chart_of_accounts=1000, starts 5/6, distinct) → left join gl_group_mapping + gl_group_dim → 430 rows (137 matched, 293 unmatched SAP codes)",
     "Jakkarit", "26-May-26", "Done"),
]


# ── helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color="000000", bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size, name="Calibri")

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── sharepoint helpers ────────────────────────────────────────────────────────
def _get_token():
    app = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        client_credential=CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Token error: {result.get('error_description', result)}")
    return result["access_token"]


def _resolve_drive(token: str) -> tuple[str, str]:
    """Return (site_id, drive_id)."""
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_SITE}:/teams/{SHAREPOINT_TEAM}",
        headers=headers,
    )
    r.raise_for_status()
    site_id = r.json()["id"]

    r = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive", headers=headers)
    r.raise_for_status()
    drive_id = r.json()["id"]
    return site_id, drive_id


def download_from_sharepoint(filename: str) -> bytes | None:
    """Download existing file from SharePoint. Returns None if not found."""
    try:
        token = _get_token()
        _, drive_id = _resolve_drive(token)
        url = (
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
            f"/root:/{SHAREPOINT_FOLDER}/{filename}:/content"
        )
        r = requests.get(url, headers={"Authorization": f"Bearer {token}"})
        if r.status_code == 404:
            print("  No existing file on SharePoint — will create fresh.")
            return None
        r.raise_for_status()
        print(f"  Downloaded {len(r.content):,} bytes from SharePoint.")
        return r.content
    except Exception as e:
        print(f"  Could not download from SharePoint: {e}")
        return None


def upload_to_sharepoint(file_bytes: bytes, filename: str) -> str | None:
    token = _get_token()
    _, drive_id = _resolve_drive(token)
    url = (
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}"
        f"/root:/{SHAREPOINT_FOLDER}/{filename}:/content"
    )
    r = requests.put(
        url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"},
        data=file_bytes,
    )
    if r.status_code in (200, 201):
        web_url = r.json().get("webUrl", "")
        print(f"  Uploaded: {web_url}")
        return web_url
    print(f"  Upload failed {r.status_code}: {r.text[:300]}")
    return None


# ── merge logic ───────────────────────────────────────────────────────────────
def read_excel_tasks(file_bytes: bytes) -> list[tuple]:
    """Read task rows from Excel bytes → list of 6-tuples."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or (row[0] is None and row[1] is None):
            continue
        rows.append((
            str(row[0] or ""),  # Action type
            str(row[1] or ""),  # Action
            str(row[2] or ""),  # Remark
            str(row[3] or ""),  # PIC
            str(row[4] or ""),  # Cut-off
            str(row[5] or ""),  # Status
        ))
    return rows


def merge_tasks(local_rows: list[tuple], sp_rows: list[tuple]) -> tuple[list[tuple], list[tuple]]:
    """
    Upsert local_rows into sp_rows by Action name (index 1).
    Returns (merged_rows, colleague_only_rows).
    Order: SP rows first (overwritten by local where matched), then new local-only rows appended.
    """
    local_by_action = {r[1]: r for r in local_rows}
    sp_actions_seen = set()

    merged = []
    for sp_row in sp_rows:
        action = sp_row[1]
        sp_actions_seen.add(action)
        merged.append(local_by_action.get(action, sp_row))  # local wins on match

    # Append new rows that exist locally but not on SP
    for row in local_rows:
        if row[1] not in sp_actions_seen:
            merged.append(row)

    local_actions = {r[1] for r in local_rows}
    colleague_rows = [r for r in sp_rows if r[1] not in local_actions]
    return merged, colleague_rows


# ── excel builder ─────────────────────────────────────────────────────────────
def build_excel(rows: list[tuple]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.sheet_view.showGridLines = False

    col_widths = [18, 38, 60, 11, 12, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["Action type", "Action", "Remark", "PIC", "Cut-off", "Status", "Data source"]
    ws.row_dimensions[1].height = 22
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _fill(C["header_bg"])
        cell.font = _font(C["header_txt"], bold=True, size=10)
        cell.alignment = _center()
        cell.border = _border()

    for row_idx, (atype, action, remark, pic, cutoff, status) in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 30
        bg = TYPE_COLORS.get(atype, C["proj_bg"])
        alt = (row_idx % 2 == 0)

        for col, val in enumerate([atype, action, remark, pic, cutoff, status, ""], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _border()

            if col == 1:
                cell.fill = _fill(bg)
                cell.font = _font(C["white"], bold=True, size=9)
                cell.alignment = _center(wrap=True)
            elif col == 2:
                cell.fill = _fill(C["row_alt"] if alt else C["row_norm"])
                cell.font = _font(C["black"], bold=True, size=9)
                cell.alignment = _left(wrap=True)
            elif col == 3:
                cell.fill = _fill(C["row_alt"] if alt else C["row_norm"])
                cell.font = _font("404040", size=9)
                cell.alignment = _left(wrap=True)
            elif col == 4:
                cell.fill = _fill(C["row_alt"] if alt else C["row_norm"])
                cell.font = _font(C["black"], size=9)
                cell.alignment = _center()
            elif col == 5:
                cell.fill = _fill(C["row_alt"] if alt else C["row_norm"])
                cell.font = _font(C["black"], size=9)
                cell.alignment = _center()
            elif col == 6:
                if val == "Done":
                    cell.fill = _fill(C["done_bg"])
                    cell.font = _font(C["white"], bold=True, size=9)
                else:
                    cell.fill = _fill(C["prog_bg"])
                    cell.font = _font(C["black"], bold=True, size=9)
                cell.alignment = _center()
            elif col == 7:
                cell.fill = _fill(C["row_alt"] if alt else C["row_norm"])
                cell.font = _font("4472C4", size=9)
                cell.alignment = _center()

    ws.freeze_panes = "A2"
    return wb


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    FOLDER.mkdir(exist_ok=True)
    out_path = FOLDER / FILENAME

    # Step 1: Download SharePoint version to check for colleague tasks
    print("Step 1: Downloading current SharePoint file...")
    sp_bytes = download_from_sharepoint(FILENAME)

    # Step 2: Merge
    if sp_bytes:
        sp_rows = read_excel_tasks(sp_bytes)
        print(f"  SharePoint has {len(sp_rows)} existing rows.")
        merged, colleague_rows = merge_tasks(ROWS, sp_rows)

        if colleague_rows:
            print(f"  Preserving {len(colleague_rows)} colleague task(s) not in local ROWS:")
            for r in colleague_rows:
                print(f"    [{r[3]}] {r[1]}")
        else:
            print("  No extra colleague tasks found — local ROWS covers all.")
    else:
        merged = list(ROWS)
        colleague_rows = []
        print(f"  Using {len(merged)} local tasks only.")

    # Step 3: Build & save locally
    print(f"\nStep 2: Building Excel ({len(merged)} rows total)...")
    wb = build_excel(merged)
    wb.save(out_path)
    print(f"  Saved locally: {out_path}")

    # Step 4: Upload
    print("\nStep 3: Uploading to SharePoint...")
    buf = io.BytesIO()
    wb.save(buf)
    try:
        url = upload_to_sharepoint(buf.getvalue(), FILENAME)
        if url:
            print(f"\nDone. SharePoint link:\n  {url}")
        else:
            print("\nExcel saved locally. Upload manually to SharePoint.")
    except Exception as e:
        print(f"\nSharePoint upload failed: {e}")
        print("Excel saved locally — upload manually.")

    # Summary
    print(f"\n--- Summary ---")
    print(f"Total rows  : {len(merged)}")
    print(f"My tasks    : {len(ROWS)}")
    print(f"Colleague   : {len(colleague_rows)}")
    if colleague_rows:
        for r in colleague_rows:
            print(f"  [{r[3]}] {r[1]} — {r[5]}")


if __name__ == "__main__":
    main()
