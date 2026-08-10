"""Append email-alert test cases to the SIT workbook on SharePoint.

Flow:
  1. Resolve the shared link -> driveItem, download the current file
  2. Append the NEW_CASES rows to sheet "2. Test Cases" (style copied from the
     last existing row), extend autofilter + dropdown validations
  3. Extend the "3. Summary" formula ranges and add the module breakdown row
  4. Save locally, then (with --upload) PUT it back, aborting if someone else
     changed the file while we were working

Usage:
  python -X utf8 setup/update_sit_test_cases.py            # build local copy only
  python -X utf8 setup/update_sit_test_cases.py --upload   # build + push to SharePoint
"""
import argparse
import base64
import io
import os
from copy import copy
from pathlib import Path

import msal
import openpyxl
import requests
from dotenv import load_dotenv
from openpyxl.worksheet.datavalidation import DataValidation

load_dotenv()

SHARE_URL = (
    "https://chememan.sharepoint.com/:x:/t/CMANDigitalTechnology/"
    "IQC3yfoY_0ExQo_fB3eD6rOoAXlplryNhGHbEroD4MlDWXA"
)
SHEET_CASES = "2. Test Cases"
SHEET_SUMMARY = "3. Summary"
MODULE = "Email Alert / Notification"
OUT_DIR = Path(__file__).parent.parent / "requirement_spec" / "4_sit"

TENANT_ID = os.getenv("ENTRA_TENANT_ID")
CLIENT_ID = os.getenv("ENTRA_CLIENT_ID")
CLIENT_SECRET = os.getenv("ENTRA_CLIENT_SECRET")

# Six cases, one per notification the app actually sends (backend/app/notifications.py):
# notify_turn, notify_reject, notify_approved, notify_step_overridden,
# notify_turn_reminder, notify_deadline_reminder.
# Tuple: (Module, Scenario, Priority, Preconditions, Steps, Test Data, Expected)
NEW_CASES = [
    (
        MODULE,
        "อีเมลแจ้งผู้อนุมัติเมื่อมีงบประมาณส่งมาถึงคิวของตน (Submit → Approver)",
        "High",
        "ระบบเปิดใช้งานการส่งอีเมลจริง (ไม่ใช่โหมดทดสอบที่ไม่ส่งออก) และผู้อนุมัติขั้นที่ 1 มีอีเมลในฐานข้อมูลพนักงาน",
        "1. ล็อกอินเป็น Filler (Pornthip)\n"
        "2. กรอกงบของฝ่าย Data Analytic แล้วกด Submit\n"
        "3. เปิดกล่องจดหมายของ Approver1 (Laddawan)\n"
        "4. คลิกลิงก์ในอีเมล",
        "ฝ่าย Data Analytic ปีงบประมาณ 2027",
        "Approver1 ได้รับอีเมลจาก CMAN_PowerBI (cmanpowerbi@chememan.com) "
        "หัวข้อ “รอการอนุมัติ งบประมาณของฝ่าย Data Analytic ปีงบประมาณ 2027”; "
        "เนื้อหาแสดง ฝ่าย / ปีงบประมาณ / ผู้ส่ง / สถานะ ครบถ้วน และลิงก์ในอีเมลเปิดเข้าหน้างบของฝ่ายและปีนั้นได้ถูกต้อง",
    ),
    (
        MODULE,
        "อีเมลแจ้งผู้ส่งเมื่องบถูกตีกลับ (Reject) พร้อมเหตุผล",
        "High",
        "มีงบสถานะรออนุมัติ และล็อกอินเป็น Approver1 ที่เปิดรายการนั้นอยู่",
        "1. กด Reject/ตีกลับ พร้อมกรอกเหตุผล\n"
        "2. เปิดกล่องจดหมายของผู้ส่ง (Pornthip)\n"
        "3. ตรวจช่อง cc ของอีเมลฉบับนั้น",
        "เหตุผล: ตัวเลขหมวด X ผิด",
        "ผู้ส่งได้รับอีเมลหัวข้อ “ถูกตีกลับ งบประมาณของฝ่าย Data Analytic ปีงบประมาณ 2027”; "
        "เนื้อหาแสดงเหตุผลที่ผู้อนุมัติกรอกไว้ครบถ้วนตรงกัน; สำเนา (cc) ถึงผู้อนุมัติขั้นที่ 1; "
        "มีลิงก์สำหรับแก้ไขและส่งใหม่",
    ),
    (
        MODULE,
        "อีเมลแจ้งผู้ส่งเมื่องบได้รับอนุมัติครบทุกขั้น (Approved)",
        "High",
        "งบผ่านการอนุมัติมาจนถึงผู้อนุมัติขั้นสุดท้าย",
        "1. อนุมัติต่อจนครบทุกขั้นของสายอนุมัติ\n"
        "2. เปิดกล่องจดหมายของผู้ส่ง\n"
        "3. นับจำนวนอีเมล “ได้รับการอนุมัติ” ที่ได้รับตลอดทั้ง flow",
        "-",
        "ผู้ส่งได้รับอีเมลหัวข้อ “ได้รับการอนุมัติ งบประมาณของฝ่าย Data Analytic ปีงบประมาณ 2027” "
        "ระบุสถานะ “อนุมัติครบทุกขั้นแล้ว”; สำเนา (cc) ถึงผู้อนุมัติขั้นที่ 1; "
        "ส่งเพียงฉบับเดียวตอนอนุมัติขั้นสุดท้าย ไม่ส่งซ้ำทุกครั้งที่มีการอนุมัติระหว่างทาง",
    ),
    (
        MODULE,
        "อีเมลแจ้งเมื่อผู้ดูแลระบบดำเนินการแทนผู้อนุมัติที่ค้าง (Step Override)",
        "Medium",
        "มีงบค้างอยู่ที่ผู้อนุมัติขั้นที่ 1 และมีบัญชีผู้ดูแลระบบ (Admin)",
        "1. ล็อกอินเป็น Admin\n"
        "2. กดปุ่มดำเนินการแทนผู้อนุมัติขั้นที่ 1\n"
        "3. เปิดกล่องจดหมายของผู้ส่ง ผู้อนุมัติที่ถูกข้าม และผู้อนุมัติคนถัดไป",
        "-",
        "ผู้ส่งได้รับอีเมลหัวข้อ “ดำเนินการแทนผู้อนุมัติ งบประมาณของฝ่าย Data Analytic ปีงบประมาณ 2027” "
        "ระบุผู้อนุมัติที่ถูกข้าม (แสดงเป็นชื่อพนักงาน ไม่ใช่รหัสพนักงาน) ผู้ดำเนินการแทน วันเวลา และสถานะปัจจุบัน; "
        "สำเนา (cc) ถึงผู้อนุมัติที่ถูกข้าม; ผู้อนุมัติคนถัดไปได้รับอีเมล “รอการอนุมัติ” ตามปกติ (รวมทั้งหมด 2 ฉบับ)",
    ),
    (
        MODULE,
        "อีเมลเตือนผู้อนุมัติที่ค้างการอนุมัติเกิน 7 วัน (รวมทุกฝ่ายใน 1 ฉบับ)",
        "Medium",
        "มีงบค้างรออนุมัติจากผู้อนุมัติคนเดียวกันหลายฝ่าย และอย่างน้อย 1 ฝ่ายค้างครบ 7 วัน",
        "1. ปล่อยให้งบค้างจนครบ 7 วัน (หรือปรับวันที่ทดสอบให้ครบเกณฑ์)\n"
        "2. รันงานแจ้งเตือนตามรอบ\n"
        "3. เปิดกล่องจดหมายของผู้อนุมัติ",
        "ผู้อนุมัติ 1 คน มีงบค้าง 3 ฝ่าย ค้างมาแล้ว 9 / 7 / 2 วัน",
        "ผู้อนุมัติได้รับอีเมลฉบับเดียว หัวข้อ “[เตือน] มีงบประมาณ 3 ฝ่ายรอการอนุมัติจากท่าน”; "
        "ในตารางแสดงครบทุกฝ่ายที่รออยู่ (รวมฝ่ายที่ยังค้างไม่ถึง 7 วัน) พร้อมจำนวนวันที่ค้างและลิงก์แยกรายฝ่าย; "
        "เตือนซ้ำทุก 7 วันจนกว่าผู้อนุมัติจะดำเนินการ",
    ),
    (
        MODULE,
        "อีเมลเตือนผู้กรอกงบที่ยังไม่ได้ส่งงบก่อนวันปิดรับ (รวมทุกฝ่ายใน 1 ฉบับ)",
        "Medium",
        "ตั้งวันเริ่มเตือนและวันปิดรับงบไว้แล้ว และ Filler ยังมีฝ่ายที่สถานะร่าง (Draft) หรือถูกตีกลับ",
        "1. ตั้งวันปิดรับให้ปัจจุบันอยู่ในช่วงเตือน\n"
        "2. รันงานแจ้งเตือนตามรอบ\n"
        "3. เปิดกล่องจดหมายของ Filler และตรวจช่อง cc",
        "Filler 1 คน รับผิดชอบ 2 ฝ่ายที่ยังไม่ส่งงบ",
        "Filler ได้รับอีเมลฉบับเดียว หัวข้อ “แจ้งเตือน: ยังไม่ได้ส่งงบประมาณ 2 ฝ่าย ปีงบประมาณ 2027”; "
        "แสดงรายชื่อฝ่ายที่ยังไม่ส่งครบทุกฝ่าย พร้อมลิงก์กรอกงบแยกรายฝ่ายและวันกำหนดปิดรับ; "
        "สำเนา (cc) ถึงผู้อนุมัติขั้นที่ 1 ของ Filler คนนั้น",
    ),
]


# ── SharePoint helpers ───────────────────────────────────────────────────────
def _get_token() -> str:
    app = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        client_credential=CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Token error: {result.get('error_description', result)}")
    return result["access_token"]


def _share_id(url: str) -> str:
    b64 = base64.b64encode(url.encode("utf-8")).decode("ascii")
    return "u!" + b64.rstrip("=").replace("/", "_").replace("+", "-")


def resolve_item(token: str) -> dict:
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{_share_id(SHARE_URL)}/driveItem",
        headers={"Authorization": f"Bearer {token}"},
    )
    r.raise_for_status()
    item = r.json()
    return {
        "id": item["id"],
        "name": item["name"],
        "driveId": item["parentReference"]["driveId"],
        "lastModifiedDateTime": item.get("lastModifiedDateTime"),
        "lastModifiedBy": (item.get("lastModifiedBy") or {}).get("user", {}).get("displayName"),
        "webUrl": item.get("webUrl"),
    }


def download(token: str, item: dict) -> bytes:
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{item['driveId']}/items/{item['id']}/content",
        headers={"Authorization": f"Bearer {token}"},
    )
    r.raise_for_status()
    return r.content


def upload(token: str, item: dict, data: bytes) -> str:
    r = requests.put(
        f"https://graph.microsoft.com/v1.0/drives/{item['driveId']}/items/{item['id']}/content",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"},
        data=data,
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Upload failed {r.status_code}: {r.text[:400]}")
    return r.json().get("webUrl", "")


# ── workbook edit ────────────────────────────────────────────────────────────
def _extend_sqref(dv: DataValidation, old_last: int, new_last: int) -> None:
    """Rewrite a validation range that ended at `old_last` to end at `new_last`."""
    dv.sqref = str(dv.sqref).replace(f"{old_last}", f"{new_last}")


def edit_workbook(raw: bytes) -> tuple[bytes, int, int]:
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb[SHEET_CASES]

    last_row = ws.max_row
    template_row = last_row
    last_no = int(ws.cell(row=last_row, column=1).value)
    first_new_row = last_row + 1

    for offset, case in enumerate(NEW_CASES):
        r = first_new_row + offset
        module, scenario, priority, pre, steps, data, expected = case
        values = {
            1: last_no + offset + 1,
            2: f"TC-{last_no + offset + 1:03d}",
            3: module,
            4: scenario,
            5: priority,
            6: pre,
            7: steps,
            8: data,
            9: expected,
            11: "Not Run",
        }
        for col in range(1, ws.max_column + 1):
            src = ws.cell(row=template_row, column=col)
            dst = ws.cell(row=r, column=col)
            dst._style = copy(src._style)
            dst.value = values.get(col)
        ws.row_dimensions[r].height = ws.row_dimensions[template_row].height

    new_last = first_new_row + len(NEW_CASES) - 1
    ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=ws.max_column).column_letter}{new_last}"
    for dv in ws.data_validations.dataValidation:
        _extend_sqref(dv, last_row, new_last)

    # Summary: stretch every formula range and add the module breakdown row.
    sm = wb[SHEET_SUMMARY]
    for row in sm.iter_rows(min_row=1, max_row=sm.max_row, max_col=sm.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = cell.value.replace(f"${last_row}", f"${new_last}")

    breakdown_last = max(
        (c.row for c in sm["B"] if isinstance(c.value, str) and c.value and c.row >= 16
         and sm.cell(row=c.row, column=3).value not in (None, "")),
        default=24,
    )
    target = breakdown_last + 1
    if sm.cell(row=target, column=2).value not in (None, ""):
        raise RuntimeError(f"Summary row {target} is not empty — refusing to overwrite")
    quoted = f"'{SHEET_CASES}'"
    formulas = {
        2: MODULE,
        3: f'=COUNTIF({quoted}!$C$4:$C${new_last},"{MODULE}")',
        4: f'=COUNTIFS({quoted}!$C$4:$C${new_last},"{MODULE}",{quoted}!$K$4:$K${new_last},"Pass")',
        5: f'=COUNTIFS({quoted}!$C$4:$C${new_last},"{MODULE}",{quoted}!$K$4:$K${new_last},"Fail")',
    }
    for col, val in formulas.items():
        src = sm.cell(row=breakdown_last, column=col)
        dst = sm.cell(row=target, column=col)
        dst._style = copy(src._style)
        dst.value = val

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), first_new_row, new_last


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--upload", action="store_true", help="push the edited file back to SharePoint")
    args = ap.parse_args()

    token = _get_token()
    item = resolve_item(token)
    print(f"File   : {item['name']}")
    print(f"Modified: {item['lastModifiedDateTime']} by {item['lastModifiedBy']}")

    raw = download(token, item)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    backup = OUT_DIR / f"{Path(item['name']).stem}_before_email_alert.xlsx"
    backup.write_bytes(raw)
    print(f"Backup : {backup}")

    out, first_new, new_last = edit_workbook(raw)
    local = OUT_DIR / item["name"]
    local.write_bytes(out)
    print(f"Local  : {local}  (rows {first_new}-{new_last} added, {len(NEW_CASES)} cases)")

    if not args.upload:
        print("Dry run — pass --upload to push to SharePoint.")
        return

    fresh = resolve_item(token)
    if fresh["lastModifiedDateTime"] != item["lastModifiedDateTime"]:
        raise RuntimeError(
            "ABORT: the SharePoint file changed while this script was running "
            f"({item['lastModifiedDateTime']} -> {fresh['lastModifiedDateTime']}). Re-run."
        )
    print(f"Uploaded: {upload(token, item, out)}")


if __name__ == "__main__":
    main()
