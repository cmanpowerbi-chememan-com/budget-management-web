"""Append the group-3 ("อื่นๆ" / other) countries to the `country.xlsx` master on
SharePoint — site CMANDWPRD, library "Budgeting and Management".

Why this exists (jakkaritw, 2026-08-22): the destination picker's third per-diem
tier was a FRONTEND-INVENTED single option ("อื่นๆ (Other)"); the master only ever
carried tier 1 (ในประเทศ) + tier 2 (ต่างประเทศ-อาเซียน). Users want the real country
names, so tier 3 becomes real master rows like every other country.

Group label: the DW notebook `budget_masters_lib.map_country_group` maps
"ในประเทศ"->domestic, "ต่างประเทศ-อาเซียน"->asian and **anything else -> "other"**,
so the new label needs no notebook change — but `dbo.country_group` will now hold
'other' rows, which `backend/app/reference_data._COUNTRY_GROUP_BY_NAME` must learn
(it currently SKIPS unrecognised groups, so the rows would silently not reach the
picker). Ship that backend change with this.

Safety (this edits a live admin-owned master):
- APPEND ONLY — never rewrites/reorders/deletes an existing row; a country already
  in the file is skipped and reported (idempotent, re-runnable).
- Dry-run by default; `--run` uploads.
- Upload is `If-Match` on the eTag read at download → a concurrent edit makes the
  PUT fail (412) instead of silently clobbering someone.
- Restores the package parts openpyxl drops (`docMetadata/LabelInfo.xml` = the
  Microsoft sensitivity label, `customXml/*` = the library plumbing); same helper
  contract as `setup/fill_sit_test_case_gaps.py`.
- Verifies by re-downloading and re-reading after the upload.

Run:
    python -X utf8 setup/add_country_master_other_group.py            # dry run
    python -X utf8 setup/add_country_master_other_group.py --run      # upload
"""
from __future__ import annotations

import argparse
import io
import os
import re
import sys
import zipfile
from copy import copy

import msal
import openpyxl
import requests
from dotenv import load_dotenv

SITE = "chememan.sharepoint.com:/sites/CMANDWPRD"
LIBRARY = "Budgeting and Management"
FILE = "country.xlsx"
SHEET = "Sheet1"
HEADER = ("country group", "country")

# Anything that is not one of the two known labels maps to `other` in the DW
# notebook; this text is the human-readable label the admin sees in Excel.
OTHER_GROUP_LABEL = "ต่างประเทศ-อื่นๆ"

# jakkaritw's list (2026-08-22), English per his decision. "Dubai (UAE)" keeps the
# name he used while still naming the country.
OTHER_COUNTRIES: tuple[str, ...] = (
    "United States",
    "United Kingdom",
    "Australia",
    "New Zealand",
    "Switzerland",
    "Japan",
    "South Korea",
    "Taiwan",
    "Germany",
    "Oman",
    "Singapore",
    "Argentina",
    "Italy",
    "Dubai (UAE)",
    "Saudi Arabia",
    "Norway",
)

GRAPH = "https://graph.microsoft.com/v1.0"

# --- package parts openpyxl drops (verbatim contract from fill_sit_test_case_gaps.py)
_PRESERVED_PART_PREFIXES = ("customXml/", "docMetadata/")
_CUSTOM_XML_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml"
_RELATIONSHIP_RE = re.compile(r"<Relationship\b[^>]*/>")
_OVERRIDE_RE = re.compile(r"<Override\b[^>]*/>")
_ID_ATTR_RE = re.compile(r'\sId="[^"]*"')


def _preserved_parts(raw: bytes) -> dict[str, bytes]:
    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        return {n: z.read(n) for n in z.namelist() if n.startswith(_PRESERVED_PART_PREFIXES)}


def _with_fresh_ids(elements: list[str], prefix: str) -> list[str]:
    return [
        _ID_ATTR_RE.sub("", el).replace("<Relationship", f'<Relationship Id="{prefix}{i}"', 1)
        for i, el in enumerate(elements, start=1)
    ]


def _inject(xml: str, closing_tag: str, elements: list[str]) -> str:
    return xml.replace(closing_tag, "".join(elements) + closing_tag, 1) if elements else xml


def preserve_sharepoint_parts(original: bytes, saved: bytes) -> tuple[bytes, list[str]]:
    keep = _preserved_parts(original)
    if not keep:
        return saved, []
    with zipfile.ZipFile(io.BytesIO(original)) as z:
        orig_ct = z.read("[Content_Types].xml").decode("utf-8")
        orig_pkg_rels = z.read("_rels/.rels").decode("utf-8")
        orig_wb_rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    overrides = [
        el for el in _OVERRIDE_RE.findall(orig_ct)
        if any(f'PartName="/{prefix}' in el for prefix in _PRESERVED_PART_PREFIXES)
    ]
    pkg_rels = _with_fresh_ids(
        [el for el in _RELATIONSHIP_RE.findall(orig_pkg_rels) if "docMetadata/" in el], "rIdSPpkg"
    )
    wb_rels = _with_fresh_ids(
        [el for el in _RELATIONSHIP_RE.findall(orig_wb_rels) if _CUSTOM_XML_REL_TYPE in el], "rIdSPcx"
    )

    out = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(saved)) as src,
        zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst,
    ):
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "[Content_Types].xml":
                data = _inject(data.decode("utf-8"), "</Types>", overrides).encode("utf-8")
            elif item.filename == "_rels/.rels":
                data = _inject(data.decode("utf-8"), "</Relationships>", pkg_rels).encode("utf-8")
            elif item.filename == "xl/_rels/workbook.xml.rels":
                data = _inject(data.decode("utf-8"), "</Relationships>", wb_rels).encode("utf-8")
            dst.writestr(item, data)
        for name, data in keep.items():
            dst.writestr(name, data)
    return out.getvalue(), sorted(keep)


def graph_headers() -> dict[str, str]:
    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))
    app = msal.ConfidentialClientApplication(
        client_id=os.environ["ENTRA_CLIENT_ID"],
        client_credential=os.environ["ENTRA_CLIENT_SECRET"],
        authority=f"https://login.microsoftonline.com/{os.environ['ENTRA_TENANT_ID']}",
    )
    tok = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in tok:
        raise SystemExit(f"token failed: {str(tok.get('error_description', tok))[:300]}")
    return {"Authorization": "Bearer " + tok["access_token"]}


def resolve_file(h: dict[str, str]) -> dict:
    site_id = requests.get(f"{GRAPH}/sites/{SITE}", headers=h, timeout=60).json()["id"]
    drives = requests.get(f"{GRAPH}/sites/{site_id}/drives", headers=h, timeout=60).json()["value"]
    drive = next(d for d in drives if d["name"] == LIBRARY)
    item = requests.get(f"{GRAPH}/drives/{drive['id']}/root:/{FILE}", headers=h, timeout=60).json()
    if "id" not in item:
        raise SystemExit(f"cannot resolve {LIBRARY}/{FILE}: {str(item)[:300]}")
    return {"drive_id": drive["id"], "item_id": item["id"], "etag": item["eTag"], "size": item["size"]}


def download(h: dict[str, str], loc: dict) -> bytes:
    r = requests.get(f"{GRAPH}/drives/{loc['drive_id']}/items/{loc['item_id']}/content", headers=h, timeout=180)
    r.raise_for_status()
    return r.content


def build_updated(raw: bytes) -> tuple[bytes, list[str], list[str]]:
    """Append the missing tier-3 rows. Returns (bytes, added, skipped)."""
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb[SHEET]
    header = tuple((ws.cell(row=1, column=c).value or "").strip().lower() for c in (1, 2))
    if header != HEADER:
        raise SystemExit(f"header changed: expected {HEADER}, found {header} — STOP, do not write")

    existing = {
        str(ws.cell(row=r, column=2).value).strip(): str(ws.cell(row=r, column=1).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=2).value
    }
    added, skipped = [], []
    style_src_row = ws.max_row  # last existing data row — copy its look
    write_row = ws.max_row + 1
    for country in OTHER_COUNTRIES:
        if country in existing:
            skipped.append(f"{country} (already present, group={existing[country]!r})")
            continue
        for col, value in ((1, OTHER_GROUP_LABEL), (2, country)):
            cell = ws.cell(row=write_row, column=col, value=value)
            src = ws.cell(row=style_src_row, column=col)
            cell.font = copy(src.font)
            cell.border = copy(src.border)
            cell.fill = copy(src.fill)
            cell.alignment = copy(src.alignment)
            cell.number_format = src.number_format
        added.append(country)
        write_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    merged, restored = preserve_sharepoint_parts(raw, buf.getvalue())
    print(f"[save] restored {len(restored)} SharePoint-only part(s): {', '.join(restored) or 'none'}")
    return merged, added, skipped


def dump_rows(raw: bytes) -> list[tuple[str, str]]:
    ws = openpyxl.load_workbook(io.BytesIO(raw))[SHEET]
    return [
        (str(ws.cell(row=r, column=1).value), str(ws.cell(row=r, column=2).value))
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=2).value
    ]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--run", action="store_true", help="actually upload (default: dry run)")
    args = ap.parse_args()

    h = graph_headers()
    loc = resolve_file(h)
    original = download(h, loc)
    before = dump_rows(original)
    print(f"[read] {FILE} size={loc['size']} etag={loc['etag']} rows={len(before)}")

    updated, added, skipped = build_updated(original)
    after = dump_rows(updated)
    print(f"[plan] append {len(added)} row(s) with group label {OTHER_GROUP_LABEL!r}")
    for c in added:
        print(f"        + {OTHER_GROUP_LABEL} | {c}")
    for s in skipped:
        print(f"        = skip {s}")

    kept = [row for row in before if row in after]
    if len(kept) != len(before):
        raise SystemExit(f"ABORT: {len(before) - len(kept)} original row(s) would be lost")
    print(f"[guard] all {len(before)} original row(s) intact; new total = {len(after)}")

    if not args.run:
        print("[dry-run] nothing uploaded — re-run with --run")
        return 0
    if not added:
        print("[run] nothing to add — file already up to date, no upload")
        return 0

    url = f"{GRAPH}/drives/{loc['drive_id']}/items/{loc['item_id']}/content"
    r = requests.put(
        url,
        headers={
            **h,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "If-Match": loc["etag"],
        },
        data=updated,
        timeout=300,
    )
    if r.status_code == 412:
        raise SystemExit("ABORT: eTag mismatch — someone edited the file since download. Re-run.")
    if r.status_code == 423:
        raise SystemExit("ABORT: file locked (423) — it is open in Excel. Close it and re-run.")
    r.raise_for_status()
    print(f"[run] uploaded ok ({r.status_code})")

    live = download(h, resolve_file(h))
    live_rows = dump_rows(live)
    with zipfile.ZipFile(io.BytesIO(live)) as z:
        parts = [n for n in z.namelist() if n.startswith(_PRESERVED_PART_PREFIXES)]
    print(f"[verify] live rows={len(live_rows)} sensitivity/customXml parts={len(parts)}")
    missing = [c for c in OTHER_COUNTRIES if c not in {row[1] for row in live_rows}]
    lost = [row for row in before if row not in live_rows]
    if missing or lost or "docMetadata/LabelInfo.xml" not in parts:
        raise SystemExit(f"VERIFY FAILED missing={missing} lost={lost} label={'docMetadata/LabelInfo.xml' in parts}")
    print("[verify] PASS — all new rows present, no original row lost, sensitivity label intact")
    return 0


if __name__ == "__main__":
    sys.exit(main())
