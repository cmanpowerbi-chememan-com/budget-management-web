"""Data-parity test: web SAP actuals vs the SAP export (truth).

Per plan/sap-actuals-dw-gap-fix.md §B.1 — the DW (`gold.fact_gl_trans`, DB
`cman_dw_wh_gold`) is missing source documents because SAP daily entry-date
files were never loaded for entry-days 2026-02-28..2026-03-30 and
2026-04-30..today. The backend query (`app.sap.fetch_sap_actuals`) is
verified CORRECT; this test exists to prove the DW reload lands: it FAILS
today (expected — the bug is upstream in the DW, unfixed) and must PASS
after the pipeline reloads the missing entry-day files and rebuilds gold.

Structure:
- `parse_sap_export()` — pure offline parser for the SAP export
  (`2026_1_4.xlsx`, posting-date view = truth). No DB, Decimal-only sums.
- `TestParseSapExport` — offline unit tests for the parser.
- `test_web_actuals_match_sap_export` — live comparison (marked
  `integration`, skipped by default via pytest.ini; skips cleanly when
  backend/.env credentials or the gold DB are unreachable). On any mismatch
  it queries doc-level rows from gold.fact_gl_trans and reports which
  Document Numbers are MISSING in the DW — the report names documents, not
  just amounts. Full report: parity_report.txt (utf-8) + the assert message.

Run:
    cd backend
    python -X utf8 -m pytest tests/tests_data_sync -v            # unit only
    python -X utf8 -m pytest tests/tests_data_sync -m integration -v  # live
"""
from __future__ import annotations

import contextlib
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

EXPORT_PATH = Path(__file__).resolve().parent / "2026_1_4.xlsx"
REPORT_PATH = Path(__file__).resolve().parent / "parity_report.txt"

FISCAL_YEAR = 2026
# The export covers posting months 1-4 only (file name: 2026_1_4) — compare
# exactly those months, never 5-12 (they would be vacuous 0 == 0 or false
# positives on docs the export simply does not cover).
COMPARE_MONTHS: tuple[int, ...] = (1, 2, 3, 4)

# Header row 1 of the SAP export (exact names, order not assumed).
EXPECTED_HEADERS: tuple[str, ...] = (
    "สายงาน",
    "ฝ่าย",
    "Cost Center",
    "Account Number",
    "ชื่อไทย",
    "Short Text",
    "Grouping",
    "Posting Date",
    "year",
    "month",
    "Document Number",
    "Text",
    "Amount in Company Code Currency",
)

# Money cells are cents-exact THB; two decimals is the comparison grain.
_CENT = Decimal("0.01")


@dataclass(frozen=True)
class SapExportRow:
    """One document-level row of the SAP export (reporting grain)."""

    cost_center: str
    gl_account: str
    month: int
    document_number: str
    amount: Decimal


def parse_sap_export(
    path: Path = EXPORT_PATH,
) -> tuple[dict[tuple[str, str, int], Decimal], list[SapExportRow]]:
    """Parse the SAP export into (monthly sums, doc-level rows).

    Returns:
        sums: `{(cost_center, gl_account, month): Decimal}` — summed Amount
            in Company Code Currency per (cc, gl, posting month).
        rows: one `SapExportRow` per data row (document grain), for the
            mismatch report.

    Pure function — no DB, no app imports. Decimal only, never float, so the
    sums are cents-exact. Columns are located BY HEADER NAME (row 1), not by
    position.
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError(f"{path.name}: empty sheet — no header row")
        columns = {str(name).strip(): idx for idx, name in enumerate(header) if name is not None}
        missing = [name for name in EXPECTED_HEADERS if name not in columns]
        if missing:
            raise ValueError(f"{path.name}: header row is missing columns: {missing}")

        idx_cc = columns["Cost Center"]
        idx_gl = columns["Account Number"]
        idx_month = columns["month"]
        idx_doc = columns["Document Number"]
        idx_amount = columns["Amount in Company Code Currency"]

        sums: dict[tuple[str, str, int], Decimal] = defaultdict(Decimal)
        doc_rows: list[SapExportRow] = []
        for row in rows:
            if row is None or row[idx_cc] is None:
                continue  # tolerate trailing blank rows
            cost_center = str(row[idx_cc]).strip()
            gl_account = str(row[idx_gl]).strip()
            month = int(row[idx_month])
            document_number = str(row[idx_doc]).strip()
            amount = Decimal(str(row[idx_amount]))
            sums[(cost_center, gl_account, month)] += amount
            doc_rows.append(SapExportRow(cost_center, gl_account, month, document_number, amount))
    finally:
        workbook.close()

    return dict(sums), doc_rows


# ---------------------------------------------------------------------------
# Offline unit tests for the parser (no DB — must always PASS).
# ---------------------------------------------------------------------------


class TestParseSapExport:
    """Pinned against the SAP-export truth table in
    plan/sap-actuals-dw-gap-fix.md §1 (verified there against SAP directly)."""

    def test_monthly_sums_match_sap_truth(self) -> None:
        sums, _ = parse_sap_export()
        expected = {
            ("10AC012000", "6210900060", 1): Decimal("7979.00"),
            ("10AC012000", "6210900060", 2): Decimal("8623.00"),
            ("10AC012000", "6210900060", 3): Decimal("10345.88"),
            ("10AC012000", "6210900060", 4): Decimal("3507.00"),
            ("10AC013000", "5210600010", 1): Decimal("500.00"),
            ("10AC013000", "5210600010", 2): Decimal("500.00"),
            ("10AC013000", "5210600010", 3): Decimal("500.00"),
            ("10AC013000", "5210600010", 4): Decimal("500.00"),
        }
        assert sums == expected

    def test_sums_are_decimal_and_exact(self) -> None:
        sums, _ = parse_sap_export()
        assert sums, "export parsed to zero keys"
        for value in sums.values():
            assert isinstance(value, Decimal), "sums must be Decimal, never float"
            assert value == value.quantize(_CENT), f"{value} has sub-cent precision — not a clean THB sum"

    def test_doc_level_rows(self) -> None:
        _, rows = parse_sap_export()
        assert len(rows) == 26, f"expected 26 data rows in the export, got {len(rows)}"
        key1 = [r for r in rows if (r.cost_center, r.gl_account) == ("10AC012000", "6210900060")]
        assert len(key1) == 14, "plan §3: the SAP export has 14 doc rows for 10AC012000/6210900060"
        for row in rows:
            assert row.month in COMPARE_MONTHS, f"row outside the export's month coverage: {row}"
            assert row.document_number, "every row must carry a Document Number"

    def test_export_covers_exactly_the_two_reported_keys(self) -> None:
        sums, _ = parse_sap_export()
        keys = {(cc, gl) for cc, gl, _month in sums}
        assert keys == {("10AC012000", "6210900060"), ("10AC013000", "5210600010")}

    def test_bad_header_is_a_loud_error(self, tmp_path: Path) -> None:
        bogus = tmp_path / "bogus.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.append(["Cost Center", "Account Number"])  # missing 11 required columns
        workbook.save(bogus)
        workbook.close()
        with pytest.raises(ValueError, match="missing columns"):
            parse_sap_export(bogus)


# ---------------------------------------------------------------------------
# Live parity test — needs the gold DW. Skipped by default (integration
# marker); skips cleanly when backend/.env credentials or the DB are absent.
# ---------------------------------------------------------------------------

# Doc-level probe for the mismatch report. Same never-cut financial contract
# as app.sap.SAP_ACTUALS_SQL (ADR-0020) so the docs named here are exactly
# the docs fetch_sap_actuals would have counted.
_DW_DOCS_SQL = """
SELECT DISTINCT accounting_doc_number
FROM gold.fact_gl_trans
WHERE company_code='1000' AND doc_type<>'CO'
  AND cost_center NOT IN ('CMRY01','CMKK01','CMPB01','MNLB00','MNLB01','MNLB02','MNLB03','MNLB04')
  AND cost_center IS NOT NULL
  AND (assignment_number IS NULL OR assignment_number<>'TFRS16')
  AND fiscal_year=? AND cost_center=? AND gl_account_number=? AND period_month=?
"""


@contextlib.contextmanager
def _gold_connection_or_skip():
    """Yield an open gold-DW connection, or pytest.skip with the reason."""
    try:
        from app.config import get_settings
        from app.db import get_gold_conn
    except ImportError as exc:  # pragma: no cover - environment guard
        pytest.skip(f"backend app modules not importable: {exc}")

    settings = get_settings()
    if not all(
        [
            settings.gold_sql_server,
            settings.gold_sql_database,
            settings.entra_client_id,
            settings.entra_client_secret,
            settings.entra_tenant_id,
        ]
    ):
        pytest.skip("backend/.env gold-DW / Entra credentials absent — live parity test needs a live DB")

    ctx = get_gold_conn(settings)
    try:
        conn = ctx.__enter__()
    except Exception as exc:
        pytest.skip(f"gold DW unreachable: {exc}")
    try:
        yield conn
    finally:
        ctx.__exit__(None, None, None)


def _fetch_dw_doc_numbers(conn, cost_center: str, gl_account: str, month: int) -> set[str]:
    cursor = conn.cursor()
    try:
        cursor.execute(_DW_DOCS_SQL, FISCAL_YEAR, cost_center, gl_account, month)
        return {str(row[0]).strip() for row in cursor.fetchall()}
    finally:
        cursor.close()


@pytest.mark.integration
def test_web_actuals_match_sap_export() -> None:
    """For every (cc, gl) key in the SAP export, every covered month (1-4):
    Excel sum == fetch_sap_actuals mMM. On mismatch, name the Document
    Numbers present in the Excel but MISSING in gold.fact_gl_trans.

    EXPECTED TODAY (DW reload not yet done): FAILS; for 10AC012000/
    6210900060 the missing-doc set must be exactly {1900000370, 1110000641,
    3110001512, 3110001584, 1900000724, 1110001487} (plan §3)."""
    from app.sap import fetch_sap_actuals

    sums, doc_rows = parse_sap_export()
    keys = sorted({(cc, gl) for cc, gl, _month in sums})

    with _gold_connection_or_skip() as conn:
        # hidden_doc_periods=None (explicit): this harness tests the MIRROR
        # property of the frozen SAP_ACTUALS_SQL against the SAP export —
        # hide_document filtering (ADR-0020 amendment 2026-08-11) has its
        # own tests in test_sap.py and must not participate here.
        data = fetch_sap_actuals(conn, FISCAL_YEAR, hidden_doc_periods=None)

        mismatches: list[str] = []
        compared = 0
        for cost_center, gl_account in keys:
            actual = data.get((cost_center, gl_account))
            for month in COMPARE_MONTHS:
                expected = sums.get((cost_center, gl_account, month), Decimal("0")).quantize(_CENT)
                got = Decimal(str(round(actual[f"m{month:02d}"], 2))).quantize(_CENT) if actual else Decimal("0.00")
                compared += 1
                if got == expected:
                    continue

                excel_docs = {
                    row.document_number
                    for row in doc_rows
                    if row.cost_center == cost_center and row.gl_account == gl_account and row.month == month
                }
                dw_docs = _fetch_dw_doc_numbers(conn, cost_center, gl_account, month)
                missing_in_dw = sorted(excel_docs - dw_docs)
                extra_in_dw = sorted(dw_docs - excel_docs)
                mismatches.append(
                    f"{cost_center}/{gl_account} m{month:02d}: excel={expected} web={got} delta={got - expected}\n"
                    f"    docs MISSING in DW ({len(missing_in_dw)}): "
                    f"{', '.join(missing_in_dw) if missing_in_dw else '(none)'}\n"
                    f"    docs in DW not in Excel ({len(extra_in_dw)}): "
                    f"{', '.join(extra_in_dw) if extra_in_dw else '(none)'}"
                )

    lines = [
        f"SAP actuals parity — web (gold.fact_gl_trans via fetch_sap_actuals) vs SAP export "
        f"({EXPORT_PATH.name}), fiscal_year={FISCAL_YEAR}, months={COMPARE_MONTHS}",
        f"cells compared: {compared} | mismatches: {len(mismatches)}",
        "",
    ]
    if mismatches:
        lines.append("MISMATCH CELLS (excel = truth):")
        lines.extend(mismatches)
    else:
        lines.append("ALL CELLS MATCH — DW reload verified.")
    report = "\n".join(lines)
    REPORT_PATH.write_text(report + "\n", encoding="utf-8")

    assert not mismatches, (
        f"web actuals != SAP export in {len(mismatches)} cell(s) — the DW is missing source documents "
        f"(plan/sap-actuals-dw-gap-fix.md). Full report: {REPORT_PATH}\n\n{report}"
    )
